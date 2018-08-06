import openpyxl as xl

# Load decoded keycodes.
codeBook = xl.load_workbook(filename = "SortedDecoded.xlsx")

# Load Vista keycodes.
codeSheet = codeBook["Vista"]

# 874 rooms in B-F.
# Fill in keycode and room.
# Iterate through all keycodes and add B-F codes into sheet, in code order.
# 1866 rooms in VDS.
# Create template sheet.
book = xl.Workbook()
sheet = book.active

# Create header.
sheet["A1"] = "Keycode"
sheet["B1"] = "Room"
sheet["C1"] = "# Sparky"
sheet["D1"] = "# Room"
sheet["E1"] = "# Mail"
sheet["F1"] = "# Fob"
currentRow = 1
prevCode = 1
while currentRow <= 874:
    for codeNum in range(prevCode, 1867):
        room = str(codeSheet["B" + str(codeNum + 1)].value)
        letter = room[3]
        if letter == "B" or letter == "C" or letter == "D" or letter == "E" or letter == "F":
            sheet.append([codeNum, room])
            currentRow += 1
            # Use this to reduce amount of data searching.
            prevCode = codeNum + 1
            break
book.save("12-3Template.xlsx")

# 992 rooms in G-K.
book = xl.Workbook()
sheet = book.active

# Create header.
sheet["A1"] = "Keycode"
sheet["B1"] = "Room"
sheet["C1"] = "# Sparky"
sheet["D1"] = "# Room"
sheet["E1"] = "# Mail"
sheet["F1"] = "# Fob"
currentRow = 1
prevCode = 1
while currentRow <= 992:
    for codeNum in range(prevCode, 1867):
        room = str(codeSheet["B" + str(codeNum + 1)].value)
        letter = room[3]
        if letter == "G" or letter == "H" or letter == "I" or letter == "J" or letter == "K":
            sheet.append([codeNum, room])
            currentRow += 1
            # Use this to reduce amount of data searching.
            prevCode = codeNum + 1
            break
book.save("3-6Template.xlsx")

# 400 rooms in L.
# Can copy straight from SortedDecoded.xlsx.
codeSheet = codeBook["Villas"]
book = xl.Workbook()
sheet = book.active

# Create header.
sheet["A1"] = "Keycode"
sheet["B1"] = "Room"
sheet["C1"] = "# Sparky"
sheet["D1"] = "# Room"
sheet["E1"] = "# Mail"
sheet["F1"] = "# Fob"
for codeNum in range(1, 400):
    room = str(codeSheet["B" + str(codeNum + 1)].value)
    sheet.append([codeNum, room])
book.save("6-9Template.xlsx")