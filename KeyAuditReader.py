import openpyxl as xl

# Return amount of rooms based on building. Rows is rooms + 1.
def building_check(ltr):
    if ltr == "B" or ltr == "C" or ltr == "H":
        return 159
    elif ltr == "D" or ltr == "E" or ltr == "G":
        return 232
    elif ltr == "F":
        return 92
    elif ltr == "I":
        return 239
    elif ltr == "J":
        return 213
    else:
        return 149

def choice_check(chc, sRow):
    # Changing sparky key amount.
    if chc == "s":
        while True:
            try:
                change = int(input("How many sparky keys?\n"))
                break
            except ValueError:
                print("Not an integer.")
        auditSheet["D" + str(sRow)].value = change
    # Changing room key amount.
    elif chc == "r":
        while True:
            try:
                change = int(input("How many room keys?\n"))
                break
            except ValueError:
                print("Not an integer.")
        auditSheet["E" + str(sRow)].value = change
    # Changing mail key amount.
    elif chc == "m":
        while True:
            try:
                change = int(input("How many mail keys?\n"))
                break
            except ValueError:
                print("Not an integer.")
        auditSheet["F" + str(sRow)].value = change
    # Changing fob amount.
    elif chc == "f":
        while True:
            try:
                change = int(input("How many fobs?\n"))
                break
            except ValueError:
                print("Not an integer.")
        auditSheet["G" + str(sRow)].value = change
    # Changing clear/discrepant.
    elif chc == "c":
        change = input("Clear or discrepant?\n")
        if change == "c":
            auditSheet["H" + str(sRow)].value = "Clear"
        else:
            auditSheet["H" + str(sRow)].value = "Discrepant"
    # Returning full set of keys.
    elif chc == "set":
        auditSheet["E" + str(sRow)].value = 1
        auditSheet["F" + str(sRow)].value = 1
        auditSheet["G" + str(sRow)].value = 1
        auditSheet["H" + str(sRow)].value = "Clear"
        auditSheet["I" + str(sRow)].value = ""
    # Changing comments.
    else:
        auditSheet["I" + str(sRow)].value = input("Input comment:\n")

# Load Vista key audit.
auditBook = xl.load_workbook(filename = "VDSKeyAudit.xlsx")

# Load sorted keycodes (1-1866).
# Keycodes are stored in column A, rooms are stored in column B.
sortedBook = xl.load_workbook(filename = "SortedDecoded.xlsx")
sortedSheet = sortedBook["Vista"]

# Setup for reading audit.
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

# Reading Vista audit back to user.
for currentKey in range(1, 1867):
    # Decode building from currentKey.
    currentRoom = sortedSheet["B" + str(currentKey + 1)].value
    currentLetter = currentRoom[3]
    roomAmount = building_check(currentLetter)
    auditSheet = auditBook[currentLetter]
    # Iterate through all entries in corresponding building to search for correct room for curentKey.
    for row in range(2, roomAmount + 2):
        # Check for match.
        if currentRoom[5:] == auditSheet["A" + str(row)].value:
            selectedRow = row
            # Print all information needed.
            print("Keycode: " + str(currentKey))
            for column in columns:
                print(auditSheet[column + "1"].value + ": " + str(auditSheet[column + str(selectedRow)].value))
            break
    verify = input("Do any changes need to be made?\n")
    # Changes need to be made.
    if verify != "":
        choice = input("What field needs to be changed?\n")
        while choice != "done":
            choice_check(choice, selectedRow)
            choice = input("What field needs to be changed?\n")
    print()
auditBook.save("VDSKeyAudit.xlsx")

#Reading Villas audit back to user.
auditBook = xl.load_workbook(filename = "VAVLKeyAudit.xlsx")
auditSheet = auditBook["L"]
sortedSheet = sortedBook["Villas"]
for currentKey in range(1, 401):
    print("Keycode: " + str(currentKey))
    # Villas is in order already, so just need to iterate through
    # Print all information needed.
    for column in columns:
        print(auditSheet[column + "1"].value + ": " + str(auditSheet[column + str(currentKey + 1)].value))
    verify = input("Do any changes need to be made?\n")
    # Changes need to be made.
    if verify != "":
        choice = input("What field needs to be changed?\n")
        while choice != "done":
            choice_check(choice, currentKey + 1)
            choice = input("What field needs to be changed?\n")
    print()
auditBook.save("VAVLKeyAudit.xlsx")